import openpyxl
from django.core.mail import send_mail
from django.db.models import Q, Count, Avg, Sum, F
from django.contrib.auth import logout
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.http import HttpResponse
from django.urls import reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import generic
from agents.mixins import OrganisorAndLoginRequiredMixin
from .models import Lead, Agent, Category
from clients.models import Client
from .forms import LeadModelForm, CustomUserCreationForm, UploadLeadsForm
from rentals.models import Rental
from sales.models import Sale
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from dateutil import parser
import datetime

class SignupView(generic.CreateView):
    template_name = "registration/signup.html"
    form_class = CustomUserCreationForm

    def get_success_url(self):
        return reverse("login")

class LandingPageView(generic.TemplateView):
    template_name = "landing.html"

class LeadListView(LoginRequiredMixin, generic.ListView):
    template_name = "leads/lead_list.html"
    context_object_name = "leads"

    def get_queryset(self):
        """
        Filters the leads based on user role, search query, and follow-up date.
        """
        queryset = Lead.objects.all().order_by("name")

        # Filter by agent if the user is an agent
        if self.request.user.is_agent:
            queryset = queryset.filter(agent__user=self.request.user)

        # Get search query and follow-up date from the request
        search_query = self.request.GET.get("search", "")
        follow_up_date = self.request.GET.get("follow_up_date", "")
        status_filter = self.request.GET.get("status", None)

        # Apply search filter if search query is provided
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone_number__icontains=search_query) |
                Q(source__icontains=search_query)
            )
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Apply date filter if follow_up_date is provided
        if follow_up_date:
            queryset = queryset.filter(follow_up_date=follow_up_date)

        return queryset

    def get_context_data(self, **kwargs):
        """
        Adds active and lost leads to the context.
        """
        queryset = self.get_queryset()

        context = super().get_context_data(**kwargs)
        context['active_leads'] = queryset.exclude(status="Lead Lost").order_by("name")
        context['lost_leads'] = queryset.filter(status="Lead Lost").order_by("name")
        return context

def export_leads_to_excel(request):
    """Generate an Excel file containing all leads"""
    leads = Lead.objects.all()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="leads.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Leads"

    # Define headers
    headers = [
        "Name", "Email", "Phone Number", "Address", "Area", "City", "Status", "Contacted",
        "Follow-up Date", "Date Modified", "Agent"
    ]
    worksheet.append(headers)

    # Write lead data
    for lead in leads:
        # If agent exists, append the agent's username; otherwise, append "Unassigned"
        agent_username = lead.agent.user.username if lead.agent else "Unassigned"

        worksheet.append([
            lead.name,
            lead.email,
            lead.phone_number,
            lead.address,
            lead.area,
            lead.city,
            lead.status,
            "Yes" if lead.contacted else "No",
            lead.follow_up_date.strftime("%Y-%m-%d") if lead.follow_up_date else "N/A",
            lead.date_modified.strftime("%Y-%m-%d") if lead.date_modified else "N/A",  # Fix: use date_modified instead of follow_up_date
            agent_username,  # Append the agent's username
        ])

    workbook.save(response)
    return response

class LeadDetailView(LoginRequiredMixin, generic.DetailView):
    template_name = "leads/lead_detail.html"
    queryset = Lead.objects.all()
    context_object_name = "lead"

class LeadCreateView(LoginRequiredMixin, generic.CreateView):
    template_name = "leads/lead_create.html"
    form_class = LeadModelForm

    def form_valid(self, form):
        lead = form.save(commit=False)  # Get the Lead instance without saving immediately

        if self.request.user.is_agent:
            try:
                agent = Agent.objects.get(user=self.request.user)
                lead.agent = agent  # Set the agent to the logged-in user's agent
            except Agent.DoesNotExist:
                # Handle the case where the user is an agent but doesn't have an Agent object
                # You might want to create an Agent object for them or display an error message
                pass  # Or handle it as you see fit.

        lead.save()  # Save the Lead instance

        # The following part is the original email sending logic, adapt as needed.
        agent = lead.agent  # Get the assigned agent (this might be the same as above)

        if agent:  # Check if an agent is assigned
            subject = f"New Lead Assigned: {lead.name}"
            message = f"A new lead '{lead.name}' has been assigned to you.\n\n" \
                      f"Lead Details:\n" \
                      f"Name: {lead.name}\n" \
                      f"Email: {lead.email}\n" \
                      f"Phone: {lead.phone_number}\n" \
                      f"Address: {lead.address}\n" \
                      f"Detail: {lead.detail}\n" \
                      f"View Lead: {self.request.build_absolute_uri(reverse('leads:lead-detail', kwargs={'pk': lead.pk}))}"
            from_email = "hauspropertiescrm@gmail.com"  # Replace with your sending email
            recipient_list = [agent.user.email]  # Agent's email address

            send_mail(subject, message, from_email, recipient_list)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse("leads:lead-list")

    def get_success_url(self):
        return reverse("leads:lead-list")

class LeadUpdateView(LoginRequiredMixin, generic.UpdateView):
    template_name = "leads/lead_update.html"
    queryset = Lead.objects.all()
    form_class = LeadModelForm

    def form_valid(self, form):
        lead = form.save(commit=False)  # Get the lead instance without saving immediately
        original_agent = self.get_object().agent  # Get the original agent
        new_agent = lead.agent  # Get the new agent

        lead.save()  # Save the lead instance

        if new_agent and new_agent != original_agent:  # Check if the agent was changed
            subject = f"Lead Reassigned: {lead.name}"
            message = f"Lead '{lead.name}' has been reassigned to you.\n\n" \
                      f"Lead Details:\n" \
                      f"Name: {lead.name}\n" \
                      f"Email: {lead.email}\n" \
                      f"Phone: {lead.phone_number}\n" \
                      f"Address: {lead.address}\n" \
                      f"Detail: {lead.detail}\n" \
                      f"View Lead: {self.request.build_absolute_uri(reverse('leads:lead-detail', kwargs={'pk': lead.pk}))}"
            from_email = "hauspropertiescrm@gmail.com"  # Replace with your sending email
            recipient_list = [new_agent.user.email]  # New agent's email address

            send_mail(subject, message, from_email, recipient_list)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse("leads:lead-detail", kwargs={"pk": self.object.pk})

class LeadDeleteView(LoginRequiredMixin, generic.DeleteView):
    template_name= "leads/lead_delete.html"
    queryset = Lead.objects.all()

    def get_success_url(self):
        return reverse("leads:lead-list")

def custom_logout_view(request):
    logout(request)
    return redirect("/")

def convert_lead_to_client(request, lead_id):
    # Get the lead object
    from clients.models import Client

    lead = get_object_or_404(Lead, id=lead_id)
    
    # Create a new Client object with lead's information
    client = Client.objects.create(
        name=lead.name,
        address=lead.address,
        area=lead.area,
        city=lead.city,
        email=lead.email,
        phone_number=lead.phone_number,
        source=lead.source,
        role="Potential Client",  # You can set this to whatever makes sense, or ask for input
        contacted=lead.contacted,
        follow_up_date=lead.follow_up_date,
        birthday=None,  # If you want to add this data, you may need to adjust
        detail=lead.detail,
        agent=lead.agent,
    )

    # Delete the lead from the leads database
    lead.delete()
    
    # Redirect to the client list page or any other page you need
    return redirect('clients:client-list')  # Adjust according to your URLs

from datetime import timedelta
from django.utils import timezone
from django.db.models import Count, Sum, F
from django.shortcuts import render
from rentals.models import Rental

class DashboardView(LoginRequiredMixin, generic.TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user
        is_agent = user.is_agent
        is_organisor = user.is_organisor

        def filter_queryset(queryset):
            if is_agent:
                return queryset.filter(agent__user=user)
            return queryset

        # --- Client Statistics ---
        client_queryset = Client.objects.all()
        filtered_clients = filter_queryset(client_queryset)
        context['total_clients'] = filtered_clients.count()

        today = timezone.now().date()
        upcoming_birthdays = []
        for client in filtered_clients:
            if client.birthday:
                this_years_birthday = client.birthday.replace(year=today.year)
                days_until_birthday = (this_years_birthday - today).days

                if 0 <= days_until_birthday <= 7:
                    upcoming_birthdays.append((days_until_birthday, client))

        upcoming_birthdays.sort(key=lambda x: x[0])
        upcoming_birthdays = [client_data[1] for client_data in upcoming_birthdays]
        context['upcoming_birthdays'] = upcoming_birthdays
        context['clients_by_role'] = filtered_clients.values('role').annotate(count=Count('role'))

        # --- Lead Statistics ---
        lead_queryset = Lead.objects.all()
        filtered_leads = filter_queryset(lead_queryset)
        context['total_leads'] = filtered_leads.count()
        context['new_leads'] = filtered_leads.filter(status="New").count()
        context['to_follow_up_leads'] = filtered_leads.filter(status="To follow up").count()
        context['total_lost_leads'] = filtered_leads.filter(status="Lead Lost").count()
        context['active_leads'] = filtered_leads.exclude(status='Lead Lost').count()
        context['lost_leads'] = filtered_leads.filter(status='Lead Lost').count()
        context['leads_by_source'] = filtered_leads.values('source').annotate(count=Count('source'))

        # --- Lead Follow-up Statistics ---
        seven_days_from_now = today + timedelta(days=7)
        leads_next_7_days = filtered_leads.filter(follow_up_date__range=[today, seven_days_from_now])
        context['leads_next_7_days'] = leads_next_7_days

        leads_follow_up_today = filtered_leads.filter(follow_up_date=today)
        context['leads_follow_up_today'] = leads_follow_up_today

        # --- Rental Statistics ---
        rental_queryset = Rental.objects.all()
        filtered_rentals = filter_queryset(rental_queryset)
        context['total_rentals'] = filtered_rentals.count()

        for_rental_rentals = filtered_rentals.exclude(status__in=["Managing", "Tenant Placed", "Closed", "Deal Lost"])
        context['for_rental_rentals_count'] = for_rental_rentals.count()
        context['total_for_rent_value'] = for_rental_rentals.aggregate(total_value=Sum('listing_price'))['total_value'] or 0

        rented_rentals = filtered_rentals.filter(status__in=["Managing", "Tenant Placed"])
        context['total_rented_value'] = rented_rentals.aggregate(total_rented_price=Sum('rental_price'))['total_rented_price'] or 0

        context['managing_rentals_count'] = filtered_rentals.filter(status="Managing").count()
        context['closed_rentals_count'] = filtered_rentals.filter(status="Closed").count()
        context['lost_rentals_count'] = filtered_rentals.filter(status="Deal Lost").count()

        days_on_market_values = [r.days_on_market() for r in filtered_rentals if r.days_on_market() is not None]
        context['average_days_on_market_rentals'] = sum(days_on_market_values) / len(days_on_market_values) if days_on_market_values else 0

        price_drop_values = [float(r.price_drop_percentage()) for r in filtered_rentals if r.price_drop_percentage() is not None]
        context['average_price_drop_rentals'] = sum(price_drop_values) / len(price_drop_values) if price_drop_values else 0
        
        #total value listed and rented rentals.
        context['total_listed_rentals_value'] = filtered_rentals.aggregate(total_listed_value=Sum('listing_price'))['total_listed_value'] or 0
        context['total_rented_rentals_value'] = filtered_rentals.filter(status__in=['Managing', 'Tenant Placed']).aggregate(total_rented_value=Sum('rental_price'))['total_rented_value'] or 0

        # --- Sale Statistics ---
        sale_queryset = Sale.objects.all()
        filtered_sales = filter_queryset(sale_queryset)
        context['total_sales'] = filtered_sales.count()

        for_sale_sales = filtered_sales.exclude(status__in=["Deal Won", "Deal Lost"])
        context['for_sale_sales_count'] = for_sale_sales.count()
        context['total_for_sale_value'] = for_sale_sales.aggregate(total_value=Sum('listing_price'))['total_value'] or 0

        sold_sales = filtered_sales.filter(status="Deal Won")
        context['sold_sales_count'] = sold_sales.count()
        context['total_sold_value'] = sold_sales.aggregate(total_sold_price=Sum('sale_price'))['total_sold_price'] or 0

        days_on_market_values_sales = [s.days_on_market() for s in filtered_sales if s.days_on_market() is not None]
        context['average_days_on_market_sales'] = sum(days_on_market_values_sales) / len(days_on_market_values_sales) if days_on_market_values_sales else 0

        price_drop_values_sales = [float(s.price_drop_percentage()) if s.price_drop_percentage() is not None else 0 for s in filtered_sales]
        context['average_price_drop_sales'] = sum(price_drop_values_sales) / len(price_drop_values_sales) if price_drop_values_sales else 0

        #total value listed and sold sales.
        context['total_listed_sales_value'] = filtered_sales.aggregate(total_listed_value=Sum('listing_price'))['total_listed_value'] or 0
        context['total_sold_sales_value'] = filtered_sales.filter(status='Deal Won').aggregate(total_sold_value=Sum('sale_price'))['total_sold_value'] or 0

        # --- Expiring Rentals ---
        expiring_rentals_30_days = filtered_rentals.filter(
            rental_expiry__lte=today + timedelta(days=30),
            rental_expiry__gte=today
        )
        context['expiring_rentals_30_days'] = expiring_rentals_30_days

        # --- Expiring Rental Mandates ---
        rental_mandates = filtered_rentals.filter(
            mandate_expiry__lte=today + timedelta(days=30),
            mandate_expiry__gte=today
        )
        context['expiring_rental_mandates_30_days'] = rental_mandates 

        # --- Expiring Sales Mandates ---
        expiring_sales_30_days = filtered_sales.filter(
            mandate_expiry__lte=today + timedelta(days=30), # Corrected field name
            mandate_expiry__gte=today # Corrected field name
        )
        context['expiring_sales_30_days'] = expiring_sales_30_days

        return context

        # --- Agent Statistics ---
        if is_organisor:
            context['leads_by_agent'] = Lead.objects.all().values('agent__user__username').annotate(count=Count('agent'))
            context['clients_by_agent'] = Client.objects.all().values('agent__user__username').annotate(count=Count('agent'))
            context['rentals_by_agent'] = Rental.objects.all().values('agent__user__username').annotate(count=Count('agent'))
            context['sales_by_agent'] = Sale.objects.all().values('agent__user__username').annotate(count=Count('agent'))

            context['sales_value_by_agent'] = Sale.objects.all().values('agent__user__username').annotate(
                total_listed=Sum('listing_price'),
                total_sold=Sum('sale_price', filter=F('status') == 'Deal Won')
            )
            context['rentals_value_by_agent'] = Rental.objects.all().values('agent__user__username').annotate(
                total_listed=Sum('listing_price'),
                total_rented=Sum('rental_price', filter=F('status__in') == ['Managing', 'Tenant Placed'])
            )

            context['sales_value_by_area'] = Sale.objects.all().values('area').annotate(
                total_listed=Sum('listing_price'),
                total_sold=Sum('sale_price', filter=F('status') == 'Deal Won')
            )
            context['rentals_value_by_area'] = Rental.objects.all().values('area').annotate(
                total_listed=Sum('listing_price'),
                total_rented=Sum('rental_price', filter=F('status__in') == ['Managing', 'Tenant Placed'])
            )

        return context

def upload_leads(request):
    if request.method == "POST":
        form = UploadLeadsForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active

                headers = [cell.value for cell in worksheet[1]]

                # Get column indexes, defaulting to None if not found
                name_index = headers.index("Name") if "Name" in headers else None
                email_index = headers.index("Email") if "Email" in headers else None
                phone_number_index = headers.index("Phone Number") if "Phone Number" in headers else None
                address_index = headers.index("Address") if "Address" in headers else None
                area_index = headers.index("Area") if "Area" in headers else None
                city_index = headers.index("City") if "City" in headers else None
                status_index = headers.index("Status") if "Status" in headers else None
                follow_up_date_index = headers.index("Follow Up Date") if "Follow Up Date" in headers else None
                agent_index = headers.index("Agent") if "Agent" in headers else None
                detail_index = headers.index("Detail") if "Detail" in headers else None
                source_index = headers.index("Source") if "Source" in headers else None

                for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    try:
                        name = row[name_index].value if name_index is not None else ""
                        email = row[email_index].value if email_index is not None else ""
                        phone_number = row[phone_number_index].value if phone_number_index is not None else ""
                        address = row[address_index].value if address_index is not None else ""
                        area = row[area_index].value if area_index is not None else ""
                        city = row[city_index].value if city_index is not None else ""
                        status = row[status_index].value if status_index is not None else "New Lead"  # Default to "New" or whatever is applicable.
                        detail = row[detail_index].value if detail_index is not None else ""
                        source = row[source_index].value if source_index is not None else ""

                        # Parse follow-up date
                        follow_up_date_value = row[follow_up_date_index].value if follow_up_date_index is not None else None
                        follow_up_date = None  # Initialize follow_up_date outside the if block

                        if follow_up_date_value:
                            try:
                                if isinstance(follow_up_date_value, datetime.datetime):
                                    follow_up_date = follow_up_date_value.date()
                                else:
                                    follow_up_date = parser.parse(str(follow_up_date_value)).date()
                            except (ValueError, TypeError):
                                messages.error(request, f"Invalid date format in row {row_num} for Follow Up Date. Date ignored.")
                                # Optionally, you could log this error for debugging.
                                # import logging
                                # logging.error(f"Invalid date format in row {row_num}: {follow_up_date_value}")

                        # Agent assignment
                        agent_username = row[agent_index].value if agent_index is not None else None
                        agent = None

                        if agent_username:
                            try:
                                agent = Agent.objects.get(user__username=agent_username)
                            except Agent.DoesNotExist:
                                messages.error(request, f"Agent '{agent_username}' not found in row {row_num}. Lead not assigned.")
                                continue  # Skip the row if the agent doesn't exist
                        else:
                            agent = Agent.objects.filter(user=request.user).first()

                        # Create Lead object
                        Lead.objects.create(
                            name=name,
                            email=email,
                            phone_number=phone_number,
                            address=address,
                            area=area,
                            city=city,
                            status=status,
                            follow_up_date=follow_up_date,
                            agent=agent,
                            detail=detail,
                            source=source,
                        )

                    except Exception as e:
                        messages.error(request, f"Error processing row {row_num}: {e}")
                        continue

                messages.success(request, "Leads uploaded successfully!")
                return redirect("leads:lead-list")

            except openpyxl.utils.exceptions.InvalidFileException:
                messages.error(request, "Invalid Excel file format.")  # More specific error
            except Exception as e:
                messages.error(request, f"Error reading file: {e}")
        else:
            messages.error(request, "Invalid form submission.")
    else:
        form = UploadLeadsForm()
    return render(request, "leads/upload_leads.html", {"form": form})