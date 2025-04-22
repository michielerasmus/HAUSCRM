import openpyxl
from django.core.mail import send_mail
from django.db.models import Q
from django.contrib.auth import logout
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.http import HttpResponse
from django.urls import reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import generic
from agents.mixins import OrganisorAndLoginRequiredMixin
from .models import Client
from .forms import ClientModelForm, CustomUserCreationForm, UploadClientsForm
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from dateutil import parser
import datetime
import logging
from leads.models import Agent

class ClientListView(LoginRequiredMixin, generic.ListView):
    template_name = "clients/client_list.html"
    context_object_name = "clients"

    def get_queryset(self):
        queryset = Client.objects.all()

        if self.request.user.is_agent:
            queryset = queryset.filter(agent__user=self.request.user)

        search_query = self.request.GET.get("search", "")
        role_filter = self.request.GET.get("role", "")
        follow_up_date = self.request.GET.get("follow_up_date", "")

        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone_number__icontains=search_query)
            )

        if role_filter:
            queryset = queryset.filter(role=role_filter)

        if follow_up_date:
            queryset = queryset.filter(follow_up_date=follow_up_date)

        return queryset.order_by("name")

def export_clients_to_excel(request):
    """Generate an Excel file containing all leads"""
    clients = Client.objects.all()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="clients.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Clients"

    # Define headers
    headers = [
        "Name", "Birthday", "Email", "Phone Number", "Address", "Area", "City", "Role", "Contacted", "Follow-up Date", "Date Modified", "Agent"
    ]
    worksheet.append(headers)

    # Write lead data
    for client in clients:
        worksheet.append([
            client.name,
            client.birthday.strftime("%Y-%m-%d") if client.birthday else "N/A",
            client.email,
            client.phone_number,
            client.address,
            client.area,
            client.city,
            client.role,
            "Yes" if client.contacted else "No",
            client.follow_up_date.strftime("%Y-%m-%d") if client.follow_up_date else "N/A",
            client.date_modified.strftime("%Y-%m-%d") if client.follow_up_date else "N/A",
            f"{client.agent.user.first_name} {client.agent.user.last_name}" if client.agent else "Unassigned",
        ])

    workbook.save(response)
    return response

class ClientDetailView(LoginRequiredMixin, generic.DetailView):
    template_name = "clients/client_detail.html"
    queryset = Client.objects.all()
    context_object_name = "client"

class ClientCreateView(LoginRequiredMixin, generic.CreateView):
    template_name = "clients/client_create.html"
    form_class = ClientModelForm

    def get_success_url(self):
        return reverse("clients:client-list")

class ClientUpdateView(LoginRequiredMixin, generic.UpdateView):
    template_name = "clients/client_update.html"
    queryset = Client.objects.all()
    form_class = ClientModelForm

    def get_success_url(self):
        return reverse("clients:client-detail", kwargs={"pk": self.object.pk})

class ClientDeleteView(LoginRequiredMixin, generic.DeleteView):
    template_name= "clients/client_delete.html"
    queryset = Client.objects.all()

    def get_success_url(self):
        return reverse("clients:client-list")

def upload_clients(request):
    if request.method == "POST":
        form = UploadClientsForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active

                headers = [cell.value for cell in worksheet[1]]

                # Get column indexes, defaulting to None if not found
                name_index = headers.index("Name") if "Name" in headers else None
                email_index = headers.index("Email") if "Email" in headers else None
                phone_number_index = headers.index("Phone Number") if "Phone Number" in headers else None
                address_index = headers.index("Address") if "Address" in headers else None
                area_index = headers.index("Area") if "Area" in headers else None
                city_index = headers.index("City") if "City" in headers else None
                role_index = headers.index("Role") if "Role" in headers else None
                follow_up_date_index = headers.index("Follow Up Date") if "Follow Up Date" in headers else None
                birthday_index = headers.index("Birthday") if "Birthday" in headers else None
                agent_index = headers.index("Agent") if "Agent" in headers else None
                detail_index = headers.index("Detail") if "Detail" in headers else None
                source_index = headers.index("Source") if "Source" in headers else None

                for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    try:
                        name = row[name_index].value if name_index is not None else ""
                        email = row[email_index].value if email_index is not None else ""
                        phone_number = row[phone_number_index].value if phone_number_index is not None else ""
                        address = row[address_index].value if address_index is not None else ""
                        area = row[area_index].value if area_index is not None else ""
                        city = row[city_index].value if city_index is not None else ""
                        role = row[role_index].value if role_index is not None else "Seller"
                        detail = row[detail_index].value if detail_index is not None else ""
                        source = row[source_index].value if source_index is not None else ""

                        # Parse follow-up date
                        follow_up_date_value = row[follow_up_date_index].value if follow_up_date_index is not None else None
                        follow_up_date = None

                        if follow_up_date_value:
                            try:
                                if isinstance(follow_up_date_value, datetime.datetime):
                                    follow_up_date = follow_up_date_value.date()
                                else:
                                    follow_up_date = parser.parse(str(follow_up_date_value)).date()
                            except (ValueError, TypeError):
                                messages.error(request, f"Invalid date format in row {row_num} for Follow Up Date. Date ignored.")
                                logging.error(f"Invalid date format in row {row_num}: {follow_up_date_value}")
                                continue

                        # Parse birthday
                        birthday_value = row[birthday_index].value if birthday_index is not None else None
                        birthday = None

                        if birthday_value:
                            try:
                                if isinstance(birthday_value, datetime.datetime):
                                    birthday = birthday_value.date()
                                else:
                                    birthday = parser.parse(str(birthday_value)).date()
                            except (ValueError, TypeError):
                                messages.error(request, f"Invalid date format in row {row_num} for Birthday. Date ignored.")
                                logging.error(f"Invalid date format in row {row_num}: {birthday_value}")
                                continue

                        # Agent assignment
                        agent_username = row[agent_index].value if agent_index is not None else None
                        agent = None

                        if agent_username:
                            try:
                                agent = Agent.objects.get(user__username=agent_username)
                            except Agent.DoesNotExist:
                                messages.error(request, f"Agent '{agent_username}' not found in row {row_num}. Lead not assigned.")
                                logging.error(f"Agent '{agent_username}' not found in row {row_num}.")
                                continue
                        else:
                            agent = Agent.objects.filter(user=request.user).first()

                        # Create Client object
                        Client.objects.create(
                            name=name,
                            email=email,
                            phone_number=phone_number,
                            address=address,
                            area=area,
                            city=city,
                            role=role,
                            follow_up_date=follow_up_date,
                            agent=agent,
                            detail=detail,
                            source=source,
                            birthday=birthday,
                        )
                        logging.debug(f"Row {row_num} processed successfully")

                    except Exception as e:
                        messages.error(request, f"Error processing row {row_num}: {e}")
                        logging.error(f"Error processing row {row_num}: {e}", exc_info=True)
                        continue

                messages.success(request, "Clients uploaded successfully!")
                return redirect("clients:client-list")

            except openpyxl.utils.exceptions.InvalidFileException:
                messages.error(request, "Invalid Excel file format.")
                logging.error("Invalid Excel file format.")
            except Exception as e:
                messages.error(request, f"Error reading file: {e}")
                logging.error(f"Error reading file: {e}", exc_info=True)
        else:
            messages.error(request, "Invalid form submission.")
            logging.error("Invalid form submission.")
    else:
        form = UploadClientsForm()
    return render(request, "clients/upload_clients.html", {"form": form})